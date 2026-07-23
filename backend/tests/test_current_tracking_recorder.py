import csv
import json
import tempfile
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from backend.app.services.current_tracking_recorder import (
    CurrentTrackingRecordingManager,
)


def tracking_point(
    elapsed_s: float,
    current_a: float | None,
    *,
    valid: bool = True,
    invalid_reason: str = "",
) -> dict:
    return {
        "elapsed_s": elapsed_s,
        "valid": valid,
        "invalid_reason": invalid_reason,
        "estimated_current_a": current_a,
        "current_sigma_a": 0.004,
        "left_frequency_hz": 2.865e9 + elapsed_s,
        "right_frequency_hz": 2.875e9 + elapsed_s,
        "splitting_hz": 10e6,
        "delta_f_sigma_hz": 1500.0,
        "common_mode_hz": 2.87e9 + elapsed_s,
        "left_state": "LOCKED",
        "right_state": "LOCKED",
        "left_error_hz": 120.0,
        "right_error_hz": -80.0,
        "left_quality": 0.95,
        "right_quality": 0.96,
        "relock_count": 0,
        "lost_lock_count": 0,
        "timing": {"measured_update_rate_hz": 5.0},
    }


class CurrentTrackingRecorderTests(unittest.TestCase):
    def test_one_second_aggregation_and_excel_export(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            manager = CurrentTrackingRecordingManager(
                Path(temporary_directory) / "records"
            )
            started = manager.start(
                interval_s=1.0,
                label="2A_13h",
                request_snapshot={"record_interval_s": 1.0, "kp": 0.45},
                device_snapshot={"microwave": {"idn": "KEYSIGHT,N5181B"}},
            )
            session_id = started["session_id"]

            self.assertFalse(manager.record_point(tracking_point(0.0, 2.0))[0])
            self.assertFalse(manager.record_point(tracking_point(0.4, 2.004))[0])
            wrote, status = manager.record_point(tracking_point(1.1, 1.996))
            self.assertTrue(wrote)
            self.assertEqual(status["rows_written"], 1)

            manager.record_point(
                tracking_point(
                    1.5,
                    None,
                    valid=False,
                    invalid_reason="left_stale",
                )
            )
            manager.record_point(tracking_point(2.2, 2.008))
            finished = manager.finish("completed")

            self.assertEqual(finished["rows_written"], 2)
            self.assertEqual(finished["valid_rows"], 1)
            self.assertTrue(finished["download_available"])
            self.assertTrue(Path(finished["csv_path"]).exists())
            self.assertTrue(Path(finished["xlsx_path"]).exists())

            with Path(finished["csv_path"]).open(
                "r",
                newline="",
                encoding="utf-8-sig",
            ) as handle:
                rows = list(csv.DictReader(handle))
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0]["sample_count"], "3")
            self.assertAlmostEqual(float(rows[0]["current_a"]), 2.0)
            self.assertEqual(rows[1]["all_samples_valid"], "false")
            self.assertEqual(rows[1]["invalid_reason"], "left_stale")

            metadata = json.loads(
                Path(finished["csv_path"])
                .with_name("metadata.json")
                .read_text(encoding="utf-8")
            )
            self.assertEqual(metadata["session_id"], session_id)
            self.assertEqual(metadata["status"], "completed")

            workbook = load_workbook(
                finished["xlsx_path"],
                read_only=True,
                data_only=False,
            )
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["Summary", "Data", "Parameters"],
                )
                data = workbook["Data"]
                data_rows = list(data.iter_rows())
                self.assertEqual(len(data_rows), 3)
                self.assertIsInstance(data_rows[1][0].value, datetime)
                self.assertIsInstance(data_rows[1][3].value, (int, float))
                self.assertEqual(data_rows[2][12].value, False)
                summary = workbook["Summary"]
                formulas = [
                    cell.value
                    for row in summary.iter_rows()
                    for cell in row
                    if isinstance(cell.value, str) and cell.value.startswith("=")
                ]
                self.assertTrue(formulas)
            finally:
                workbook.close()

    def test_active_recording_can_export_a_snapshot(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            manager = CurrentTrackingRecordingManager(Path(temporary_directory))
            started = manager.start(
                interval_s=1.0,
                label="live",
                request_snapshot={},
                device_snapshot={},
            )
            manager.record_point(tracking_point(0.0, 1.0))
            manager.record_point(tracking_point(1.0, 1.1))
            path, status = manager.export(started["session_id"])

            self.assertTrue(path.exists())
            self.assertEqual(status["status"], "recording")
            self.assertEqual(status["rows_written"], 1)
            manager.finish("cancelled")

    def test_long_tracking_gap_is_not_blended_into_one_interval(self) -> None:
        with tempfile.TemporaryDirectory() as temporary_directory:
            manager = CurrentTrackingRecordingManager(Path(temporary_directory))
            manager.start(
                interval_s=1.0,
                label="gap",
                request_snapshot={},
                device_snapshot={},
            )
            manager.record_point(tracking_point(0.0, 1.0))
            wrote, _ = manager.record_point(tracking_point(5.0, 3.0))
            self.assertTrue(wrote)
            finished = manager.finish("completed")

            with Path(finished["csv_path"]).open(
                "r",
                newline="",
                encoding="utf-8-sig",
            ) as handle:
                rows = list(csv.DictReader(handle))
            self.assertEqual(len(rows), 2)
            self.assertAlmostEqual(float(rows[0]["current_a"]), 1.0)
            self.assertAlmostEqual(float(rows[1]["current_a"]), 3.0)


if __name__ == "__main__":
    unittest.main()
